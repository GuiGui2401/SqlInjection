import argparse
import requests
import re
import openpyxl
from fpdf import FPDF

def check_sql_injection(url, input):
    response = requests.get(url, params={"input": input})

    # Recherche d'une erreur de syntaxe SQL
    if re.search(r"Error in your SQL syntax", response.text):
        return True

    # Recherche d'une fuite de données
    if re.search(r"'", response.text):
        return True

    return False

def create_excel_report(url, input):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "URL"
    sheet["B1"] = "Input"
    sheet["C1"] = "Résultat"
    sheet.append([url, input, "Faille SQL injection trouvée !"])
    workbook.save("rapport_sql_injection.xlsx")

def create_pdf_report(url, input):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, "Rapport de test des failles SQL injection", 1, 1, "C")
    pdf.cell(200, 10, "URL", 1, 0, "L")
    pdf.cell(200, 10, "Input", 1, 0, "L")
    pdf.cell(200, 10, "Résultat", 1, 0, "L")
    pdf.cell(200, 10, url, 1, 1, "L")
    pdf.cell(200, 10, input, 1, 1, "L")
    pdf.cell(200, 10, "Aucune faille SQL injection trouvée.", 1, 1, "L")
    pdf.output("rapport_sql_injection.pdf")

def main():
    parser = argparse.ArgumentParser(description="Test de failles SQL injection")
    parser.add_argument("url", help="URL à tester")
    parser.add_argument("input_file", help="Fichier texte contenant les requêtes SQL à tester (une par ligne)")
    parser.add_argument("-t", "--report-type", choices=["excel", "pdf"], default="excel", help="Type de rapport (excel ou pdf)")
    args = parser.parse_args()

    with open(args.input_file, "r") as file:
        queries = file.readlines()

    for query in queries:
        query = query.strip()  # Supprimer les espaces et les sauts de ligne supplémentaires
        if check_sql_injection(args.url, query):
            if args.report_type == "excel":
                create_excel_report(args.url, query)
            elif args.report_type == "pdf":
                create_pdf_report(args.url, query)
            print(f"Faille SQL injection trouvée pour la requête : {query}")

if __name__ == "__main__":
    main()
